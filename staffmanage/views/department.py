# -*- coding: UTF8 -*- #
"""
@filename:department.py
@author:Ajie
@time:2024-07-24
"""
from django.shortcuts import render, redirect
from django.http import HttpResponse

from openpyxl import load_workbook

from staffmanage import models
from staffmanage.utils.pagination import Pagination


# Create your views here.
def depart_list(request):
    """部门列表"""

    # 在数据库获取所有的部门列表
    # 【对象，对象，对象 ...】
    queryset = models.Department.objects.all()

    # 使用封装好的类进行分页
    page_obj = Pagination(request, queryset)

    context = {
        'departments': page_obj.page_queryset,
        'page_string': page_obj.HTML(),
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户POST请求提交的数据
    depart_name = request.POST.get('departName')

    if depart_name == '':
        return redirect('/depart/add')

    # 保存到数据库
    models.Department.objects.create(title=depart_name)

    # 重定向
    return redirect('/depart/list')


def depart_delete(request):
    """删除部门"""

    # http://127.0.0.1:8000/depart/delete/?depart_id=1
    depart_id = request.GET.get('depart_id')
    if depart_id == '':
        return redirect('/depart/list')

    models.Department.objects.filter(id=depart_id).delete()
    return redirect('/depart/list')


def depart_edit(request, depart_id):
    """修改部门"""

    if request.method == "GET":
        # 根据depart_id，获取其部门名 [obj,]
        row_obj = models.Department.objects.filter(id=depart_id).first()
        return render(request, 'depart_edit.html', {'depart_obj': row_obj})

    title = request.POST.get('departName')
    if title == '':
        return redirect('/depart/list/')

    models.Department.objects.filter(id=depart_id).update(title=title)

    return redirect('/depart/list/')


def depart_upload(request):
    """批量上传部门(Excel)"""
    # 1 获取用户上传的文件对象
    file_obj = request.FILES.get('file')
    if not file_obj:
        return redirect('/depart/list/')

    # 2 对象传递给openpyxl，读取内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    # 3 循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
